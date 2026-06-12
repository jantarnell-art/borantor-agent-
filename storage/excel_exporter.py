"""Export all data to Excel with one sheet per data category."""

from __future__ import annotations

import logging
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import EXCEL_PATH, BINDING_PERIODS
from storage import database as db

logger = logging.getLogger(__name__)

# Colour palette
HEADER_FILL = PatternFill("solid", fgColor="1F497D")   # dark blue
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_ROW_FILL = PatternFill("solid", fgColor="DCE6F1")  # light blue
WARN_FILL = PatternFill("solid", fgColor="FFE699")      # amber
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")      # green
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")       # red


def _header_row(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def export_all() -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _sheet_list_rates(wb)
    _sheet_avg_rates(wb)
    _sheet_reference_rates(wb)
    _sheet_my_offers(wb)
    _sheet_analysis(wb)
    _sheet_warnings(wb)

    wb.save(EXCEL_PATH)
    logger.info("Excel rapport sparad: %s", EXCEL_PATH)


# ── Sheet 1: Listräntor ───────────────────────────────────────────────────────

def _sheet_list_rates(wb) -> None:
    ws = wb.create_sheet("Listräntor")
    _header_row(ws, ["Datum", "Bank", "Bindningstid", "Ränta (%)", "Källa"])
    rows = db.get_list_rates(limit=5000)
    for i, r in enumerate(rows, 1):
        ws.append([
            r["rate_date"],
            r["bank"],
            r["period_label"],
            r["rate"],
            r["source_url"] or "",
        ])
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = ALT_ROW_FILL
    _autofit(ws)


# ── Sheet 2: Snitträntor ─────────────────────────────────────────────────────

def _sheet_avg_rates(wb) -> None:
    ws = wb.create_sheet("Snitträntor")
    _header_row(ws, ["Datum", "Bank", "Bindningstid", "Snittränta (%)", "Källa"])
    with db._conn() as con:
        rows = [
            dict(r) for r in con.execute(
                "SELECT * FROM avg_rates ORDER BY rate_date DESC, bank, period_key LIMIT 5000"
            ).fetchall()
        ]
    for i, r in enumerate(rows, 1):
        ws.append([r["rate_date"], r["bank"], r["period_label"], r["rate"], r["source_url"] or ""])
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = ALT_ROW_FILL
    _autofit(ws)


# ── Sheet 3: Referensräntor ──────────────────────────────────────────────────

def _sheet_reference_rates(wb) -> None:
    ws = wb.create_sheet("Referensräntor")
    _header_row(ws, ["Datum", "Ränta", "Värde (%)", "Källa"])
    with db._conn() as con:
        rows = [
            dict(r) for r in con.execute(
                "SELECT * FROM reference_rates ORDER BY rate_date DESC, series_key LIMIT 5000"
            ).fetchall()
        ]
    for i, r in enumerate(rows, 1):
        ws.append([r["rate_date"], r["series_label"], r["rate"], r["source"] or ""])
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = ALT_ROW_FILL
    _autofit(ws)


# ── Sheet 4: Mina erbjudanden ─────────────────────────────────────────────────

def _sheet_my_offers(wb) -> None:
    ws = wb.create_sheet("Mina erbjudanden")
    _header_row(ws, [
        "Datum", "Bank", "Bindningstid", "Erbjuden ränta (%)",
        "Lånebelopp (kr)", "Rabatt mot lista (pp)", "Kommentar", "Källa",
    ])
    for i, r in enumerate(db.get_my_offers(), 1):
        ws.append([
            r["offer_date"],
            r["bank"],
            r["period_label"],
            r["offered_rate"],
            r["loan_amount"] or "",
            r["discount_vs_list"] or "",
            r["comment"] or "",
            r["source"] or "",
        ])
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = ALT_ROW_FILL
    _autofit(ws)


# ── Sheet 5: Analys ───────────────────────────────────────────────────────────

def _sheet_analysis(wb) -> None:
    from analysis.calculator import market_summary, savings_vs_alternative

    ws = wb.create_sheet("Analys")
    ws.append(["Marknadsöversikt – senaste listräntor"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])

    latest = db.get_latest_list_rates()
    summary = market_summary(latest)

    _header_row(ws, ["Bindningstid", "Lägst (%)", "Högst (%)", "Snitt (%)", "Antal banker"])
    for period_key, label in BINDING_PERIODS.items():
        s = summary.get(period_key)
        if not s:
            continue
        ws.append([label, s["min"], s["max"], s["avg"], s["count"]])
    ws.append([])

    # Per-bank latest rates table
    ws.append(["Senaste listräntor per bank"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)
    ws.append([])

    banks_sorted = sorted({r["bank"] for r in latest})
    headers = ["Bank"] + [BINDING_PERIODS[k] for k in BINDING_PERIODS]
    _header_row(ws, headers)

    for bank in banks_sorted:
        bank_rates = {r["period_key"]: r["rate"] for r in latest if r["bank"] == bank}
        row = [bank] + [bank_rates.get(k, "") for k in BINDING_PERIODS]
        ws.append(row)

    ws.append([])
    ws.append([f"Genererad: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])

    # My offer analysis
    offers = db.get_my_offers()
    if offers:
        ws.append([])
        ws.append(["Min lånesituation"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)
        ws.append([])
        _header_row(ws, [
            "Bank", "Bindningstid", "Min ränta (%)", "Listränta (%)",
            "Rabatt (pp)", "Månadskostnad (kr)*", "Årskostnad (kr)*",
        ])
        list_map = {(r["bank"], r["period_key"]): r["rate"] for r in latest}
        from analysis.calculator import monthly_cost, annual_cost, discount_vs_list

        for o in sorted(offers, key=lambda x: x["offer_date"], reverse=True):
            key = (o["bank"], o["period_key"])
            list_rate = list_map.get(key, "")
            disc = ""
            if list_rate:
                disc = discount_vs_list(o["offered_rate"], list_rate)
            loan = o["loan_amount"]
            mc = monthly_cost(loan, o["offered_rate"]) if loan else ""
            ac = annual_cost(loan, o["offered_rate"]) if loan else ""
            ws.append([
                o["bank"], o["period_label"], o["offered_rate"],
                list_rate, disc, mc, ac,
            ])

    _autofit(ws)


# ── Sheet 6: Varningar ────────────────────────────────────────────────────────

def _sheet_warnings(wb) -> None:
    ws = wb.create_sheet("Varningar")
    _header_row(ws, ["Datum", "Kategori", "Bank", "Bindningstid", "Meddelande", "Allvarlighet"])
    for r in db.get_warnings(acknowledged=False, limit=500):
        ws.append([
            r["warning_date"],
            r["category"],
            r["bank"] or "",
            r["period_key"] or "",
            r["message"],
            r["severity"],
        ])
        row_num = ws.max_row
        fill = WARN_FILL if r["severity"] == "WARNING" else ALT_ROW_FILL
        for cell in ws[row_num]:
            cell.fill = fill
    _autofit(ws)
